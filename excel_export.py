"""
Investment-Banking Grade Multi-Sheet Excel Workbook Generator
Generates NSE_BSE_Deals_Tracker.xlsx using openpyxl with exact financial formatting.
"""

import io
from datetime import date
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


# Professional Color Palette (Institutional Navy Theme)
NAVY_HEADER = "1B365D"
NAVY_SUBHEADER = "2C4D75"
WHITE = "FFFFFF"
LIGHT_GRAY = "F4F6F9"
BORDER_GRAY = "D1D5DB"
CARD_FILL = "F8FAFC"
ACCENT_BLUE = "0284C7"
GREEN_TEXT = "15803D"
RED_TEXT = "B91C1C"

FONT_NAME = "Segoe UI"


def _apply_box_border(ws, min_r, min_c, max_r, max_c):
    """Applies clean borders around a block of cells."""
    thin_border = Border(
        left=Side(style='thin', color=BORDER_GRAY),
        right=Side(style='thin', color=BORDER_GRAY),
        top=Side(style='thin', color=BORDER_GRAY),
        bottom=Side(style='thin', color=BORDER_GRAY)
    )
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(row=r, column=c).border = thin_border


def build_raw_data_sheet(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    """Creates the 'Raw Data' sheet containing the 'Deals_Data' Excel Table."""
    ws = wb.create_sheet(title="Raw Data")
    ws.views.sheetView[0].showGridLines = True

    columns = [
        ("Exchange", "exchange"),
        ("Deal Type", "deal_type"),
        ("Date", "date"),
        ("Symbol", "symbol"),
        ("Security Name", "security_name"),
        ("Client Name", "client_name"),
        ("Buy/Sell", "buy_sell"),
        ("Quantity", "quantity"),
        ("Deal Price", "price"),
        ("Deal Value", "deal_value"),
        ("Deal Value (₹ Cr)", "deal_value_cr"),
        ("Deal Date Close", "deal_date_close"),
        ("1 Trading Day Ago Close", "t1_close"),
        ("5 Trading Days Ago Close", "t5_close"),
        ("15 Trading Days Ago Close", "t15_close"),
        ("Deal Price vs Deal Date Close (%)", "deal_vs_close_pct"),
        ("Deal Date Close vs 1D Ago (%)", "close_vs_t1_pct"),
        ("Deal Date Close vs 5D Ago (%)", "close_vs_t5_pct"),
        ("Deal Date Close vs 15D Ago (%)", "close_vs_t15_pct"),
        ("Deal Price vs 15D Ago Close (%)", "deal_vs_t15_pct"),
    ]

    # Write Headers
    for c_idx, (col_title, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_title)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # Write Data
    row_idx = 2
    if not df.empty:
        for _, record in df.iterrows():
            for c_idx, (_, field_name) in enumerate(columns, 1):
                val = record.get(field_name, None)
                cell = ws.cell(row=row_idx, column=c_idx)

                if pd.isna(val) or val is None:
                    cell.value = ""
                else:
                    if field_name == "date" and isinstance(val, (date, pd.Timestamp)):
                        cell.value = val.strftime("%Y-%m-%d")
                        cell.alignment = Alignment(horizontal="center")
                    elif field_name == "quantity":
                        cell.value = float(val)
                        cell.number_format = '#,##0'
                    elif field_name in ["price", "deal_date_close", "t1_close", "t5_close", "t15_close"]:
                        cell.value = float(val)
                        cell.number_format = '₹#,##0.00'
                    elif field_name == "deal_value":
                        cell.value = float(val)
                        cell.number_format = '#,##0.00'
                    elif field_name == "deal_value_cr":
                        cell.value = float(val)
                        cell.number_format = '₹#,##0.00 "Cr"'
                    elif "pct" in field_name:
                        cell.value = float(val) / 100.0
                        cell.number_format = '0.00%'
                    else:
                        cell.value = str(val)

                cell.font = Font(name=FONT_NAME, size=9)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        # Create Table Object
        max_row = row_idx - 1
        max_col_letter = get_column_letter(len(columns))
        tab = Table(displayName="Deals_Data", ref=f"A1:{max_col_letter}{max_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tab)
    else:
        ws.cell(row=2, column=1, value="No records found for the selected period.")

    # Freeze panes below header
    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)


def build_analysis_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, stats: Dict[str, Any]) -> None:
    """Builds the primary executive-facing Analysis sheet."""
    ws = wb.create_sheet(title="Analysis")
    ws.views.sheetView[0].showGridLines = True

    # Title Block
    ws.merge_cells("A1:K1")
    title_cell = ws.cell(row=1, column=1, value="NSE & BSE DEALS TRACKER — INSTITUTIONAL DEAL FLOW ANALYSIS")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    # 1. EXECUTIVE SUMMARY SECTION
    ws.cell(row=3, column=1, value="1. EXECUTIVE SUMMARY").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)

    metrics = [
        ("Total Deals", stats.get("total_deals", 0), '#,##0'),
        ("Total Deal Value (₹ Cr)", stats.get("total_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("NSE Deal Value (₹ Cr)", stats.get("nse_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("BSE Deal Value (₹ Cr)", stats.get("bse_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("Bulk Deal Value (₹ Cr)", stats.get("bulk_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("Block Deal Value (₹ Cr)", stats.get("block_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("Buy Value (₹ Cr)", stats.get("buy_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("Sell Value (₹ Cr)", stats.get("sell_val_cr", 0.0), '₹#,##0.00 "Cr"'),
        ("Net Buy / (Sell) (₹ Cr)", stats.get("net_buy_cr", 0.0), '₹#,##0.00 "Cr"'),
    ]

    for m_idx, (m_label, m_val, m_fmt) in enumerate(metrics):
        col_c = (m_idx % 3) * 3 + 1
        row_r = 4 + (m_idx // 3) * 2

        # Card Label
        lbl_cell = ws.cell(row=row_r, column=col_c, value=m_label)
        lbl_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="475569")
        lbl_cell.fill = PatternFill(start_color=CARD_FILL, end_color=CARD_FILL, fill_type="solid")
        ws.merge_cells(start_row=row_r, start_column=col_c, end_row=row_r, end_column=col_c + 1)

        # Card Value
        val_cell = ws.cell(row=row_r + 1, column=col_c, value=m_val)
        val_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)
        val_cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        val_cell.number_format = m_fmt
        val_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_r + 1, start_column=col_c, end_row=row_r + 1, end_column=col_c + 1)

        _apply_box_border(ws, row_r, col_c, row_r + 1, col_c + 1)

    curr_row = 11

    # 2. KEY DEAL FLOW OBSERVATIONS
    ws.cell(row=curr_row, column=1, value="2. KEY DEAL FLOW OBSERVATIONS").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)
    curr_row += 1

    observations = stats.get("observations", ["No notable observations."])
    for obs in observations:
        c = ws.cell(row=curr_row, column=1, value=f"•  {obs}")
        c.font = Font(name=FONT_NAME, size=9.5)
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=10)
        curr_row += 1

    curr_row += 1

    # 3. TOP 15 STOCKS BY DEAL VALUE
    ws.cell(row=curr_row, column=1, value="3. TOP 15 STOCKS BY DEAL VALUE").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)
    curr_row += 1

    stock_cols = [
        "Rank", "Symbol", "Deals", "Total Quantity", "Buy Value (₹ Cr)", "Sell Value (₹ Cr)",
        "Total Value (₹ Cr)", "Net Buy/Sell (₹ Cr)", "Deal Date Close", "1D Δ %", "5D Δ %", "15D Δ %", "Avg Deal vs Close %"
    ]
    for c_idx, h in enumerate(stock_cols, 1):
        cell = ws.cell(row=curr_row, column=c_idx, value=h)
        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_SUBHEADER, end_color=NAVY_SUBHEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    curr_row += 1

    top_stocks = stats.get("top_stocks_df", pd.DataFrame())
    start_stock_row = curr_row
    if not top_stocks.empty:
        for r_idx, row in top_stocks.head(15).iterrows():
            ws.cell(row=curr_row, column=1, value=r_idx + 1).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=2, value=str(row.get("symbol", "")))
            ws.cell(row=curr_row, column=3, value=int(row.get("deal_count", 0))).number_format = '#,##0'
            ws.cell(row=curr_row, column=4, value=float(row.get("quantity", 0))).number_format = '#,##0'
            ws.cell(row=curr_row, column=5, value=float(row.get("buy_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=6, value=float(row.get("sell_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=7, value=float(row.get("deal_value_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=8, value=float(row.get("net_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'

            # Market context
            c_close = row.get("deal_date_close")
            if pd.notna(c_close):
                ws.cell(row=curr_row, column=9, value=float(c_close)).number_format = '₹#,##0.00'

            for p_idx, col_k in enumerate(["close_vs_t1_pct", "close_vs_t5_pct", "close_vs_t15_pct", "deal_vs_close_pct"], 10):
                v = row.get(col_k)
                if pd.notna(v):
                    p_cell = ws.cell(row=curr_row, column=p_idx, value=float(v) / 100.0)
                    p_cell.number_format = '0.00%'

            for c in range(1, len(stock_cols) + 1):
                ws.cell(row=curr_row, column=c).font = Font(name=FONT_NAME, size=9)
            curr_row += 1

        _apply_box_border(ws, start_stock_row - 1, 1, curr_row - 1, len(stock_cols))
    else:
        ws.cell(row=curr_row, column=1, value="No stock aggregate data available.")
        curr_row += 1

    curr_row += 1

    # 4. TOP 15 CLIENTS
    ws.cell(row=curr_row, column=1, value="4. TOP 15 CLIENTS BY TRANSACTION VOLUME").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)
    curr_row += 1

    client_cols = ["Rank", "Client Name", "Deal Count", "Unique Stocks", "Buy Value (₹ Cr)", "Sell Value (₹ Cr)", "Total Value (₹ Cr)", "Net Buy/Sell (₹ Cr)"]
    for c_idx, h in enumerate(client_cols, 1):
        cell = ws.cell(row=curr_row, column=c_idx, value=h)
        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_SUBHEADER, end_color=NAVY_SUBHEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    curr_row += 1

    top_clients = stats.get("top_clients_df", pd.DataFrame())
    start_client_row = curr_row
    if not top_clients.empty:
        for r_idx, row in top_clients.head(15).iterrows():
            ws.cell(row=curr_row, column=1, value=r_idx + 1).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=2, value=str(row.get("client_name", "")))
            ws.cell(row=curr_row, column=3, value=int(row.get("deal_count", 0))).number_format = '#,##0'
            ws.cell(row=curr_row, column=4, value=int(row.get("unique_stocks", 0))).number_format = '#,##0'
            ws.cell(row=curr_row, column=5, value=float(row.get("buy_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=6, value=float(row.get("sell_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=7, value=float(row.get("deal_value_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'
            ws.cell(row=curr_row, column=8, value=float(row.get("net_val_cr", 0.0))).number_format = '₹#,##0.00 "Cr"'

            for c in range(1, len(client_cols) + 1):
                ws.cell(row=curr_row, column=c).font = Font(name=FONT_NAME, size=9)
            curr_row += 1

        _apply_box_border(ws, start_client_row - 1, 1, curr_row - 1, len(client_cols))
    else:
        ws.cell(row=curr_row, column=1, value="No client aggregate data available.")
        curr_row += 1

    curr_row += 2

    # 5. METHODOLOGY & DISCLAIMERS
    ws.cell(row=curr_row, column=1, value="5. METHODOLOGY & INSTITUTIONAL DISCLAIMERS").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY_HEADER)
    curr_row += 1

    disclaimers = [
        "Deal Data Source: Official reporting feeds for Bulk Deals (≥ 0.5% equity) and Block Deals (separate trading window) from NSE and BSE.",
        "Market Price Context: Stock price histories reflect official daily closing market prices. Reference sessions (T, T-1, T-5, T-15) exclude non-trading dates and exchange holidays.",
        "Valuation Formulae: Deal Value = Quantity × Deal Price; Deal Value (₹ Cr) = Deal Value / 10,000,000; Premium/Discount = (Deal Price / Deal Date Close - 1) × 100.",
        "Regulatory Notice: Bulk/Block transaction reporting is mandatory exchange disclosure. Reported transactions alone do not establish underlying investment intent or forward expectations.",
        "Disclaimer: NSE and BSE interfaces used by this application may include unofficial/reverse-engineered endpoints and may change without notice. Verify against official publications before material investment decisions."
    ]

    for d in disclaimers:
        c = ws.cell(row=curr_row, column=1, value=f"•  {d}")
        c.font = Font(name=FONT_NAME, size=8.5, italic=True, color="475569")
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=11)
        curr_row += 1

    # Adjust widths
    for col_i in range(1, 13):
        col_letter = get_column_letter(col_i)
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30


def build_pivots_sheet(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    """
    Creates clearly labeled pivot-style summary tables using pandas aggregation
    with institutional openpyxl formatting.
    """
    ws = wb.create_sheet(title="Pivots")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="AGGREGATE TRANSACTION PIVOTS & CROSS-TABULATIONS")
    title_cell.font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    ws.row_dimensions[1].height = 30

    curr_row = 3

    if df.empty:
        ws.cell(row=curr_row, column=1, value="No records available to generate pivot summaries.")
        return

    def _write_pivot_block(title: str, pivot_df: pd.DataFrame, format_cr: bool = True):
        nonlocal curr_row
        ws.cell(row=curr_row, column=1, value=title).font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY_SUBHEADER)
        curr_row += 1

        # Write pivot headers
        headers = [pivot_df.index.name or "Dimension"] + list(pivot_df.columns)
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=curr_row, column=c_idx, value=str(h))
            cell.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=NAVY_SUBHEADER, end_color=NAVY_SUBHEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        curr_row += 1

        start_r = curr_row
        for idx_val, row in pivot_df.iterrows():
            ws.cell(row=curr_row, column=1, value=str(idx_val)).font = Font(name=FONT_NAME, size=9, bold=True)
            for c_idx, val in enumerate(row, 2):
                cell = ws.cell(row=curr_row, column=c_idx)
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = '₹#,##0.00 "Cr"' if format_cr else '#,##0'
                else:
                    cell.value = "-"
                cell.font = Font(name=FONT_NAME, size=9)
            curr_row += 1

        _apply_box_border(ws, start_r - 1, 1, curr_row - 1, len(headers))
        curr_row += 2

    # 1. Exchange x Deal Type (Value ₹ Cr)
    p1 = df.pivot_table(index="exchange", columns="deal_type", values="deal_value_cr", aggfunc="sum", fill_value=0)
    p1["Total (₹ Cr)"] = p1.sum(axis=1)
    _write_pivot_block("1. Deal Value (₹ Cr) by Exchange × Deal Type", p1)

    # 2. Deal Type x Buy/Sell (Value ₹ Cr)
    p2 = df.pivot_table(index="deal_type", columns="buy_sell", values="deal_value_cr", aggfunc="sum", fill_value=0)
    p2["Total (₹ Cr)"] = p2.sum(axis=1)
    _write_pivot_block("2. Deal Value (₹ Cr) by Deal Type × Buy/Sell", p2)

    # 3. Daily Deal Value Trend
    p3 = df.pivot_table(index="date", columns="exchange", values="deal_value_cr", aggfunc="sum", fill_value=0)
    p3["Daily Total (₹ Cr)"] = p3.sum(axis=1)
    _write_pivot_block("3. Daily Deal Value (₹ Cr) by Exchange", p3)

    # 4. Symbol x Buy/Sell Top 10
    top_syms = df.groupby("symbol")["deal_value_cr"].sum().nlargest(10).index
    df_top_syms = df[df["symbol"].isin(top_syms)]
    p4 = df_top_syms.pivot_table(index="symbol", columns="buy_sell", values="deal_value_cr", aggfunc="sum", fill_value=0)
    p4["Total (₹ Cr)"] = p4.sum(axis=1)
    _write_pivot_block("4. Top 10 Stocks × Buy/Sell Breakdown (₹ Cr)", p4)

    for c_i in range(1, 8):
        ws.column_dimensions[get_column_letter(c_i)].width = 22


def generate_excel_workbook(deals_df: pd.DataFrame, stats: Dict[str, Any]) -> io.BytesIO:
    """
    Builds and packages the institutional workbook into an in-memory BytesIO buffer.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build the three primary worksheets
    build_raw_data_sheet(wb, deals_df)
    build_analysis_sheet(wb, deals_df, stats)
    build_pivots_sheet(wb, deals_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
